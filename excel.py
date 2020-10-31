from openpyxl import Workbook, load_workbook
from config import RESULT_FILENAME, COLUMN_SIZE, TITLE_COLUMNS, NEW_SHEET
from os.path import exists


def setup_new_workbook() -> Workbook:
    """
        Попереднє налаштування вивідного Ексель файлу, а саме:
        збільшення колонок;
        додаємо назви колонок до стовпчиків;
        та інше
    :return:
    """
    if exists(RESULT_FILENAME):
        workbook: Workbook = load_workbook(RESULT_FILENAME)
        workbook.create_sheet(NEW_SHEET)  # створюємо в документі нову таблицю
        sheet = workbook.worksheets[-1]  # кожен раз при новому записі записуємо в нову таблицю
    else:
        workbook: Workbook = Workbook()
        sheet = workbook.active
    sheet.title = 'Статьи'
    for index, column in enumerate(COLUMN_SIZE):     # регулюємо ширину колонок
        sheet.column_dimensions[chr(65 + index)].width = column
    sheet.append(TITLE_COLUMNS)     # додаємо назви колонок
    return workbook


def write_to_excel(articles: list) -> bool:
    """
        запис даних статей в файл
    :param articles:
    :return:
    """
    workbook: Workbook = setup_new_workbook()
    for article in articles:
        workbook.worksheets[-1].append(article)     # пишемо в об'єкт Ексель файлу по одній статті
    try:
        workbook.save(RESULT_FILENAME)
    except PermissionError:     # якщо файл використовується його не можна перезаписати
        return False
    else:
        return True


if __name__ == '__main__':
    pass